from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal
from sqlalchemy.orm import Session
from .analytics import get_overdue_loans, get_recovery_metrics, get_all_clients_with_loans, get_portfolio_summary
import io

def generate_recovery_report(db: Session) -> bytes:
    """Generate Excel report for overdue loans."""
    overdue = get_overdue_loans(db)
    metrics = get_recovery_metrics(db)
    
    wb = Workbook()
    
    # --- Sheet 1: Overdue Loans ---
    ws = wb.active
    ws.title = "Overdue Loans"
    
    # Header
    headers = ["Loan ID", "Client Name", "Phone", "Principal (UGX)", "EMI (UGX)", "Paid (UGX)", "Outstanding (UGX)", "Days Overdue", "Status"]
    ws.append(headers)
    
    # Apply header styling
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for loan in overdue:
        ws.append([
            loan["loan_id"],
            loan["customer_name"],
            loan["customer_phone"],
            float(loan["principal"]),
            float(loan["monthly_emi"]),
            float(loan["total_paid"]),
            float(loan["outstanding_balance"]),
            loan["days_overdue"],
            loan["status"]
        ])
    
    # Auto-adjust column widths
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # --- Sheet 2: Recovery Summary ---
    ws2 = wb.create_sheet("Recovery Summary")
    ws2.append(["Metric", "Value (UGX)"])
    ws2.append(["Total Expected", float(metrics["total_expected"])])
    ws2.append(["Total Collected", float(metrics["total_collected"])])
    ws2.append(["Total Principal", float(metrics["total_principal"])])
    ws2.append(["Repayment Rate", f"{metrics['repayment_rate']:.2f}%"])
    ws2.append(["Portfolio Yield", f"{metrics['portfolio_yield']:.2f}%"])
    
    for row in range(1, 7):
        ws2.row_dimensions[row].height = 25
    
    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def generate_financial_statements(db: Session) -> bytes:
    """Generate Profit & Loss and Balance Sheet."""
    summary = get_portfolio_summary(db)
    metrics = get_recovery_metrics(db)
    
    wb = Workbook()
    
    # --- Sheet 1: Profit & Loss ---
    ws = wb.active
    ws.title = "Profit & Loss"
    
    ws.append(["PROFIT & LOSS STATEMENT"])
    ws.merge_cells("A1:C1")
    ws["A1"].font = Font(size=16, bold=True)
    
    ws.append([])
    ws.append(["INCOME"])
    ws["A3"].font = Font(bold=True)
    ws.append(["Interest Income", float(metrics["total_collected"])])
    
    ws.append([])
    ws.append(["EXPENSES"])
    ws["A6"].font = Font(bold=True)
    ws.append(["Loan Loss Provision", "0"])
    
    ws.append([])
    ws.append(["NET PROFIT"])
    ws["A9"].font = Font(bold=True)
    ws.append(["Net Profit", float(metrics["total_collected"])])
    
    # --- Sheet 2: Balance Sheet ---
    ws2 = wb.create_sheet("Balance Sheet")
    ws2.append(["BALANCE SHEET"])
    ws2.merge_cells("A1:C1")
    ws2["A1"].font = Font(size=16, bold=True)
    
    ws2.append([])
    ws2.append(["ASSETS"])
    ws2["A3"].font = Font(bold=True)
    ws2.append(["Gross Loan Portfolio", float(summary["total_principal"])])
    ws2.append(["Cash", "0"])
    ws2.append(["Total Assets", float(summary["total_principal"])])
    
    ws2.append([])
    ws2.append(["LIABILITIES & EQUITY"])
    ws2["A8"].font = Font(bold=True)
    ws2.append(["Equity", float(summary["total_principal"])])
    ws2.append(["Total Liabilities & Equity", float(summary["total_principal"])])
    
    # Auto-adjust column widths
    for ws in [ws, ws2]:
        for col in range(1, 4):
            ws.column_dimensions[get_column_letter(col)].width = 30
    
    # Save
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def generate_portfolio_report(db: Session) -> bytes:
    """Generate portfolio report with expected vs actual returns."""
    clients = get_all_clients_with_loans(db)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio Report"
    
    # Header
    headers = ["Client", "Phone", "Principal (UGX)", "Rate %", "Tenure", "EMI (UGX)", "Expected (UGX)", "Collected (UGX)", "Outstanding (UGX)", "Status"]
    ws.append(headers)
    
    # Styling
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    total_expected = Decimal(0)
    total_collected = Decimal(0)
    
    for client in clients:
        ws.append([
            client["client_name"],
            client["client_phone"],
            float(client["principal"]),
            float(client["interest_rate"]),
            client["tenure"],
            float(client["emi"]),
            float(client["total_expected"]),
            float(client["total_collected"]),
            float(client["outstanding"]),
            client["status"]
        ])
        total_expected += client["total_expected"]
        total_collected += client["total_collected"]
    
    # Summary row
    ws.append([])
    ws.append(["TOTAL", "", "", "", "", "", float(total_expected), float(total_collected), float(total_expected - total_collected), ""])
    
    # Auto-adjust columns
    for col in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()